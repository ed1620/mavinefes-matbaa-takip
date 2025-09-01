from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from flask_mail import Mail, Message
from flask_socketio import SocketIO, emit
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_caching import Cache
import sqlite3
import os
import sys
import logging
from logging.handlers import RotatingFileHandler

from datetime import datetime, timedelta
import bcrypt
import random
import string
import json
import uuid
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from dotenv import load_dotenv

# Environment variables yükle
load_dotenv()

app = Flask(__name__)

# Production static files için whitenoise
if os.environ.get('FLASK_ENV') == 'production':
    try:
        from whitenoise import WhiteNoise
        app.wsgi_app = WhiteNoise(app.wsgi_app, root='static/')
        app.wsgi_app.add_files('static/', prefix='static/')
    except ImportError:
        pass  # whitenoise yoksa devam et

app.secret_key = os.environ.get('SECRET_KEY', 'matbaa_takip_2025_secret_key')

# CSRF Protection (temporarily disabled for testing)
# csrf = CSRFProtect(app)

# Rate Limiting - Production/Development
if os.environ.get('FLASK_ENV') == 'production' and os.environ.get('REDIS_URL'):
    # Redis varsa kullan
    limiter = Limiter(
        get_remote_address,
        app=app,
        default_limits=["200 per day", "50 per hour"],
        storage_uri=os.environ.get('REDIS_URL')
    )
else:
    # Redis yoksa memory storage kullan
    limiter = Limiter(
        get_remote_address,
        app=app,
        default_limits=["200 per day", "50 per hour"]
    )

# Cache Configuration - Production/Development
if os.environ.get('FLASK_ENV') == 'production' and os.environ.get('REDIS_URL'):
    # Redis varsa kullan
    cache_config = {
        'CACHE_TYPE': 'redis',
        'CACHE_REDIS_URL': os.environ.get('REDIS_URL'),
        'CACHE_DEFAULT_TIMEOUT': 300
    }
else:
    # Redis yoksa simple cache kullan
    cache_config = {
        'CACHE_TYPE': 'simple',
        'CACHE_DEFAULT_TIMEOUT': 300
    }

cache = Cache(app, config=cache_config)

# Logging Configuration - Windows uyumlu
if not app.debug:
    try:
        if not os.path.exists('logs'):
            os.mkdir('logs')
        file_handler = RotatingFileHandler('logs/matbaa_takip.log', maxBytes=10240, backupCount=5)
        file_handler.setFormatter(logging.Formatter(
            '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
        ))
        file_handler.setLevel(logging.INFO)
        app.logger.addHandler(file_handler)
        app.logger.setLevel(logging.INFO)
        app.logger.info('Matbaa Takip Sistemi başlatılıyor')
    except Exception as e:
        # Windows'ta logging hatası olursa console'a yaz
        print(f'Logging hatası: {e}')
        app.logger.setLevel(logging.INFO)
        app.logger.info('Matbaa Takip Sistemi başlatılıyor (console mode)')
else:
    app.logger.setLevel(logging.INFO)
    app.logger.info('Matbaa Takip Sistemi başlatılıyor (debug mode)')

# Production/Development config
if os.environ.get('FLASK_ENV') == 'production':
    app.config['DEBUG'] = False
    # Production database URL (PostgreSQL için)
    DATABASE_URL = os.environ.get('DATABASE_URL')
    if DATABASE_URL and DATABASE_URL.startswith('postgres://'):
        DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)
    
    # Production secret key
    app.secret_key = os.environ.get('SECRET_KEY', 'production_secret_key_change_this')
else:
    app.config['DEBUG'] = True
    DATABASE_URL = None
    app.secret_key = os.environ.get('SECRET_KEY', '')  # ❌ Default key kaldırıldı!

# Email konfigürasyonu - Environment variables ile
app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', 'True').lower() == 'true'
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', 'eren1121623@gmail.com')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', '')  # ❌ Şifre kaldırıldı!
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_DEFAULT_SENDER', 'eren1121623@gmail.com')

# E-posta gönderme aktif/pasif (test için aktif)
EMAIL_ENABLED = True

mail = Mail(app)

# SocketIO konfigürasyonu - Production/Development
if os.environ.get('FLASK_ENV') == 'production':
    socketio = SocketIO(
        app, 
        cors_allowed_origins="*",
        async_mode='threading',  # eventlet yerine threading
        logger=True,
        engineio_logger=True
    )
else:
    socketio = SocketIO(app, cors_allowed_origins="*")

# SQLite veritabanı dosyası (development için)
DATABASE = os.environ.get('DATABASE_PATH', 'matbaa_takip.db')

# Session timeout configuration
app.permanent_session_lifetime = timedelta(hours=2)

# Security headers - Production için
if os.environ.get('FLASK_ENV') == 'production':
    @app.after_request
    def add_security_headers(response):
        response.headers['X-Content-Type-Options'] = 'nosniff'
        response.headers['X-Frame-Options'] = 'DENY'
        response.headers['X-XSS-Protection'] = '1; mode=block'
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
        return response

def init_db():
    """Veritabanını oluştur ve tabloları hazırla"""
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    
    # Books tablosu - İyileştirilmiş
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS books (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            author_name TEXT NOT NULL,
            order_quantity INTEGER NOT NULL,
            size TEXT NOT NULL,
            status TEXT DEFAULT 'Hazırlanıyor',
            track_code TEXT UNIQUE NOT NULL,
            customer_email TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # İndeksler ekle - Performans iyileştirmesi
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_track_code ON books(track_code)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_status ON books(status)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_created_at ON books(created_at)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_customer_email ON books(customer_email)')
    
    # Users tablosu (admin girişi için)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT DEFAULT 'admin',
            is_active BOOLEAN DEFAULT 1,
            last_login TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Contact messages tablosu (iletişim formu için)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS contact_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT NOT NULL,
            message TEXT NOT NULL,
            is_read BOOLEAN DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    

    
    # Eğer is_read kolonu yoksa ekle (mevcut tablolar için)
    try:
        cursor.execute('ALTER TABLE contact_messages ADD COLUMN is_read BOOLEAN DEFAULT 0')
    except:
        pass  # Kolon zaten varsa hata vermez
    
    # Admin kullanıcısını güncelle veya oluştur - Environment variables ile
    admin_username = os.environ.get('ADMIN_USERNAME', 'admin')
    admin_password = os.environ.get('ADMIN_PASSWORD', 'admin123')
    
    cursor.execute('DELETE FROM users WHERE username IN (?, ?)', ('admin', admin_username))
    hashed_password = hash_password(admin_password)
    cursor.execute('INSERT INTO users (username, password) VALUES (?, ?)', (admin_username, hashed_password))
    
    conn.commit()
    conn.close()
    app.logger.info('Veritabanı başarıyla başlatıldı')

def get_db_connection():
    """Veritabanı bağlantısı oluştur - SQLite/PostgreSQL desteği"""
    try:
        if os.environ.get('FLASK_ENV') == 'production' and os.environ.get('DATABASE_URL'):
            # PostgreSQL production
            import psycopg2
            from psycopg2.extras import RealDictCursor
            
            conn = psycopg2.connect(
                os.environ.get('DATABASE_URL'),
                cursor_factory=RealDictCursor
            )
        else:
            # SQLite development
            conn = sqlite3.connect(DATABASE, timeout=10.0)
            conn.row_factory = sqlite3.Row  # Dict-like access
            conn.execute('PRAGMA foreign_keys = ON')  # Foreign key desteği
        
        return conn
    except Exception as e:
        app.logger.error(f'Veritabanı bağlantı hatası: {e}')
        return None

def hash_password(password):
    """Bcrypt ile güvenli şifre hash'leme"""
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode('utf-8'), salt).decode('utf-8')

def verify_password(password, hashed):
    """Şifre doğrulama"""
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))



def generate_track_code():
    """Benzersiz takip kodu oluştur - İyileştirilmiş"""
    # Farklı format seçenekleri
    formats = [
        # Format 1: 3 harf + 6 rakam (örn: ABC123456)
        lambda: ''.join(random.choices(string.ascii_uppercase, k=3)) + ''.join(random.choices(string.digits, k=6)),
        # Format 2: 2 harf + 4 rakam + 2 harf (örn: AB1234CD)
        lambda: ''.join(random.choices(string.ascii_uppercase, k=2)) + ''.join(random.choices(string.digits, k=4)) + ''.join(random.choices(string.ascii_uppercase, k=2)),
        # Format 3: 4 rakam + 4 harf (örn: 1234ABCD)
        lambda: ''.join(random.choices(string.digits, k=4)) + ''.join(random.choices(string.ascii_uppercase, k=4)),
        # Format 4: 2 harf + 6 rakam (örn: TR123456)
        lambda: ''.join(random.choices(string.ascii_uppercase, k=2)) + ''.join(random.choices(string.digits, k=6))
    ]
    
    # Rastgele bir format seç
    selected_format = random.choice(formats)
    return selected_format()

def is_track_code_unique(track_code):
    """Takip kodunun benzersiz olup olmadığını kontrol et"""
    if not track_code:
        return False
    
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM books WHERE track_code = ?", (track_code,))
            count = cursor.fetchone()[0]
            return count == 0
        except Exception as e:
            app.logger.error(f"Takip kodu kontrol hatası: {e}")
            return False
        finally:
            conn.close()
    return False

def get_unique_track_code():
    """Benzersiz takip kodu oluştur ve döndür - İyileştirilmiş"""
    max_attempts = 20
    
    for attempt in range(max_attempts):
        track_code = generate_track_code()
        if is_track_code_unique(track_code):
            app.logger.info(f"Benzersiz takip kodu oluşturuldu (deneme {attempt + 1}): {track_code}")
            return track_code
    
    # Eğer maksimum denemede benzersiz kod bulunamazsa, timestamp ekle
    app.logger.warning("Benzersiz kod bulunamadı, timestamp ekleniyor...")
    timestamp = datetime.now().strftime('%H%M%S')
    base_code = generate_track_code()
    final_code = f"{base_code}{timestamp}"
    
    # Son kontrol
    if is_track_code_unique(final_code):
        app.logger.info(f"Timestamp ile benzersiz kod oluşturuldu: {final_code}")
        return final_code
    else:
        # Son çare: UUID kullan
        unique_id = str(uuid.uuid4())[:8].upper()
        final_code = f"TRK{unique_id}"
        app.logger.info(f"UUID ile benzersiz kod oluşturuldu: {final_code}")
        return final_code

def update_book_timestamp(book_id):
    """Kitap güncelleme zamanını güncelle"""
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("UPDATE books SET updated_at = CURRENT_TIMESTAMP WHERE id = ?", (book_id,))
            conn.commit()
        except Exception as e:
            app.logger.error(f"Timestamp güncelleme hatası: {e}")
        finally:
            conn.close()



def validate_email(email):
    """E-posta adresi validasyonu"""
    import re
    if not email:
        return False
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def validate_book_data(title, author_name, order_quantity, size):
    """Kitap verisi validasyonu"""
    errors = []
    
    if not title or len(title.strip()) < 2:
        errors.append('Kitap adı en az 2 karakter olmalıdır.')
    
    if not author_name or len(author_name.strip()) < 2:
        errors.append('Yazar adı en az 2 karakter olmalıdır.')
    
    try:
        qty = int(order_quantity)
        if qty <= 0 or qty > 10000:
            errors.append('Sipariş adedi 1-10000 arasında olmalıdır.')
    except (ValueError, TypeError):
        errors.append('Geçerli bir sipariş adedi giriniz.')
    
    if not size or len(size.strip()) < 2:
        errors.append('Kitap boyutu en az 2 karakter olmalıdır.')
    
    return errors

def sanitize_input(text):
    """Kullanıcı girdilerini temizle"""
    if not text:
        return ''
    
    # HTML karakterlerini escape et
    import html
    text = html.escape(text.strip())
    
    # Maksimum uzunluk kontrolü
    if len(text) > 1000:
        text = text[:1000]
    
    return text



def send_email_notification(to_email, subject, body, track_code=None):
    """Email bildirimi gönder - İyileştirilmiş"""
    if not EMAIL_ENABLED:
        app.logger.info(f"E-posta gönderme pasif: {subject} -> {to_email}")
        return True
    
    if not app.config['MAIL_USERNAME'] or not app.config['MAIL_PASSWORD']:
        app.logger.warning("E-posta konfigürasyonu eksik")
        return False
    
    try:
        msg = Message(
            subject=subject,
            sender=app.config['MAIL_USERNAME'],
            recipients=[to_email]
        )
        msg.html = body
        
        # Logo ekle
        try:
            logo_path = os.path.join(app.static_folder, 'img', 'logo.png')
            if os.path.exists(logo_path):
                with open(logo_path, 'rb') as logo_file:
                    msg.attach('logo.png', 'image/png', logo_file.read())
        except Exception as logo_error:
            app.logger.warning(f"Logo ekleme hatası: {logo_error}")
        
        mail.send(msg)
        app.logger.info(f"E-posta başarıyla gönderildi: {subject} -> {to_email}")
        return True
    except Exception as e:
        app.logger.error(f"Email gönderme hatası: {e}")
        return False

def send_track_code_email(book_data, customer_email=None):
    """Takip kodu oluşturulduğunda email gönder"""
    subject = f"Takip Kodunuz Oluşturuldu - {book_data['title']}"
    
    body = f"""
    <!DOCTYPE html>
    <html lang="tr">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Takip Kodu Oluşturuldu</title>
    </head>
    <body style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 0; padding: 0; background-color: #f5f5f5;">
        <div style="max-width: 600px; margin: 20px auto; background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 10px rgba(0, 0, 0, 0.1);">
            
            <!-- Header -->
            <div style="background-color: #007bff; padding: 30px; text-align: center;">
                <h1 style="color: white; margin: 0; font-size: 24px; font-weight: 600;">Mavi Nefes Matbaa</h1>
            </div>
            
            <!-- Content -->
            <div style="padding: 30px;">
                <h2 style="color: #333; margin-bottom: 20px; font-size: 20px; font-weight: 600;">Siparişiniz Başarıyla Alındı</h2>
                
                <p style="color: #666; line-height: 1.6; margin-bottom: 25px; font-size: 14px;">
                    Değerli müşterimiz, kitabınız için takip kodu başarıyla oluşturuldu. 
                    Aşağıdaki bilgileri dikkatlice inceleyiniz ve takip kodunuzu güvenli bir yerde saklayınız.
                </p>
                
                <!-- Order Details -->
                <div style="background-color: #f8f9fa; border-radius: 6px; padding: 20px; margin: 20px 0; border-left: 4px solid #007bff;">
                    <h3 style="color: #333; margin: 0 0 15px 0; font-size: 16px; font-weight: 600;">Sipariş Detayları</h3>
                    
                    <table style="width: 100%; border-collapse: collapse;">
                        <tr>
                            <td style="padding: 8px 0; color: #666; font-size: 13px; width: 120px;"><strong>Kitap Adı:</strong></td>
                            <td style="padding: 8px 0; color: #333; font-size: 14px;">{book_data['title']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px 0; color: #666; font-size: 13px;"><strong>Yazar:</strong></td>
                            <td style="padding: 8px 0; color: #333; font-size: 14px;">{book_data['author_name']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px 0; color: #666; font-size: 13px;"><strong>Sipariş Adedi:</strong></td>
                            <td style="padding: 8px 0; color: #333; font-size: 14px;">{book_data['order_quantity']} adet</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px 0; color: #666; font-size: 13px;"><strong>Kitap Boyutu:</strong></td>
                            <td style="padding: 8px 0; color: #333; font-size: 14px;">{book_data['size']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px 0; color: #666; font-size: 13px;"><strong>Mevcut Durum:</strong></td>
                            <td style="padding: 8px 0; color: #28a745; font-size: 14px; font-weight: 600;">{book_data['status']}</td>
                        </tr>
                    </table>
                </div>
                
                <!-- Track Code -->
                <div style="text-align: center; margin: 25px 0;">
                    <p style="margin: 0 0 10px 0; color: #666; font-size: 13px; text-transform: uppercase; letter-spacing: 1px;"><strong>Takip Kodu</strong></p>
                    <div style="background-color: #007bff; color: white; padding: 15px 25px; border-radius: 6px; font-size: 20px; font-weight: bold; letter-spacing: 2px; display: inline-block;">
                        {book_data['track_code']}
                    </div>
                </div>
                
                <!-- How to Track -->
                <div style="background-color: #f8f9fa; border-radius: 6px; padding: 20px; margin: 20px 0;">
                    <h3 style="color: #333; margin: 0 0 15px 0; font-size: 16px; font-weight: 600;">Takip Kodunuzu Nasıl Kullanırsınız?</h3>
                    <ul style="color: #666; line-height: 1.6; margin: 0; padding-left: 20px; font-size: 14px;">
                        <li><strong>Web Sitemizi Ziyaret Edin:</strong> Takip sayfamızdan anlık bilgi alabilirsiniz</li>
                        <li><strong>Takip Kodunu Girin:</strong> Yukarıdaki takip kodunu kullanarak sorgulama yapın</li>
                    </ul>
                </div>
                
                <!-- Contact Info -->
                <div style="background-color: #e3f2fd; border-radius: 6px; padding: 20px; margin: 20px 0;">
                    <h4 style="color: #1976d2; margin: 0 0 15px 0; font-size: 16px; font-weight: 600;">İletişim Bilgileri</h4>
                    <p style="color: #333; margin: 5px 0; font-size: 14px;">
                        <strong>E-posta:</strong> siparis@mavinefes.com.tr
                    </p>
                    <p style="color: #333; margin: 5px 0; font-size: 14px;">
                        <strong>Telefon:</strong> +90 258 266 55 44
                    </p>
                    <p style="color: #333; margin: 5px 0; font-size: 14px;">
                        <strong>Adres:</strong> Mavi Nefes Yayınları, Zümrüt, Vatan Cd No:240, 20160 Denizli Merkez/Denizli
                    </p>
                </div>
            </div>
            
            <!-- Footer -->
            <div style="background-color: #6c757d; color: white; padding: 20px; text-align: center;">
                <p style="margin: 0 0 5px 0; font-size: 12px; color: rgba(255,255,255,0.8);">
                    Bu e-posta Mavi Nefes Matbaa Takip Sistemi tarafından otomatik olarak gönderilmiştir.
                </p>
                <p style="margin: 0; font-size: 11px; color: rgba(255,255,255,0.6);">
                    © 2025 Mavi Nefes Matbaa. Tüm hakları saklıdır.
                </p>
            </div>
        </div>
    </body>
    </html>
    """
    
    # Eğer müşteri email adresi verilmişse onu kullan, yoksa demo email
    to_email = customer_email if customer_email else "demo@example.com"
    return send_email_notification(to_email, subject, body, book_data['track_code'])

def send_status_update_email(book_data, new_status, customer_email=None):
    """Durum güncellemesi email'i gönder"""
    status_messages = {
        'Sipariş Alındı': 'Siparişiniz başarıyla alındı ve sisteme kaydedildi. Baskı süreciniz planlanmaya başlanacak.',
        'Bekliyor': 'Siparişiniz sırada bekliyor. Sıranız geldiğinde size bilgi verilecektir.',
        'Kontrolde': 'Kitabınız kalite kontrolünden geçiyor. Her detay titizlikle inceleniyor.',
        'Planlamada': 'Baskı planlaması yapılıyor. En uygun baskı takvimi belirleniyor.',
        'Üretimde': 'Kitabınız baskı sürecinde. Baskı işlemi devam ediyor.',
        'Hazır': 'Kitabınız teslime hazır! En kısa sürede size ulaştırılacak.'
    }
    
    status_colors = {
        'Sipariş Alındı': '#28a745',
        'Bekliyor': '#ffc107',
        'Kontrolde': '#17a2b8',
        'Planlamada': '#6f42c1',
        'Üretimde': '#fd7e14',
        'Hazır': '#dc3545'
    }
    
    status_icons = {
        'Sipariş Alındı': '📋',
        'Bekliyor': '⏳',
        'Kontrolde': '🔍',
        'Planlamada': '📅',
        'Üretimde': '🖨️',
        'Hazır': '✅'
    }
    
    subject = f"Durum Güncellendi - {book_data['title']}"
    body = f"""
    <!DOCTYPE html>
    <html lang="tr">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Durum Güncellendi</title>
    </head>
    <body style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 0; padding: 0; background-color: #f5f5f5;">
        <div style="max-width: 600px; margin: 20px auto; background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 10px rgba(0, 0, 0, 0.1);">
            
            <!-- Header -->
            <div style="background-color: #007bff; padding: 30px; text-align: center;">
                <h1 style="color: white; margin: 0; font-size: 24px; font-weight: 600;">Mavi Nefes Matbaa</h1>
            </div>
            
            <!-- Content -->
            <div style="padding: 30px;">
                <h2 style="color: #333; margin-bottom: 20px; font-size: 20px; font-weight: 600;">Kitabınızın Durumu Değişti</h2>
                
                <p style="color: #666; line-height: 1.6; margin-bottom: 25px; font-size: 14px;">
                    Değerli müşterimiz, kitabınızın durumu güncellendi. Aşağıdaki bilgileri inceleyerek 
                    güncel durumu öğrenebilirsiniz.
                </p>
                
                <!-- Book Info -->
                <div style="background-color: #f8f9fa; border-radius: 6px; padding: 20px; margin: 20px 0; border-left: 4px solid #007bff;">
                    <h3 style="color: #333; margin: 0 0 15px 0; font-size: 16px; font-weight: 600;">Kitap Bilgileri</h3>
                    
                    <table style="width: 100%; border-collapse: collapse;">
                        <tr>
                            <td style="padding: 8px 0; color: #666; font-size: 13px; width: 120px;"><strong>Kitap Adı:</strong></td>
                            <td style="padding: 8px 0; color: #333; font-size: 14px;">{book_data['title']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px 0; color: #666; font-size: 13px;"><strong>Yazar:</strong></td>
                            <td style="padding: 8px 0; color: #333; font-size: 14px;">{book_data['author_name']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px 0; color: #666; font-size: 13px;"><strong>Takip Kodu:</strong></td>
                            <td style="padding: 8px 0; color: #333; font-size: 14px; background-color: #e9ecef; padding: 4px 8px; border-radius: 4px; display: inline-block;">{book_data['track_code']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px 0; color: #666; font-size: 13px;"><strong>Sipariş Adedi:</strong></td>
                            <td style="padding: 8px 0; color: #333; font-size: 14px;">{book_data['order_quantity']} adet</td>
                        </tr>
                    </table>
                </div>
                
                <!-- Status Update -->
                <div style="text-align: center; margin: 25px 0;">
                    <p style="margin: 0 0 10px 0; color: #666; font-size: 13px; text-transform: uppercase; letter-spacing: 1px;"><strong>Yeni Durum</strong></p>
                    <div style="background-color: {status_colors.get(new_status, '#007bff')}; color: white; padding: 15px 25px; border-radius: 6px; font-size: 18px; font-weight: bold; display: inline-block;">
                        {status_icons.get(new_status, '🔄')} {new_status}
                    </div>
                </div>
                
                <!-- Status Description -->
                <div style="background-color: #f8f9fa; border-radius: 6px; padding: 20px; margin: 20px 0;">
                    <h3 style="color: #333; margin: 0 0 15px 0; font-size: 16px; font-weight: 600;">Durum Açıklaması</h3>
                    <p style="color: #666; margin: 0; line-height: 1.6; font-size: 14px;">
                        {status_messages.get(new_status, 'Durum güncellendi.')}
                    </p>
                </div>
                
                <!-- Next Steps -->
                <div style="background-color: #e8f5e8; border-radius: 6px; padding: 20px; margin: 20px 0;">
                    <h3 style="color: #333; margin: 0 0 15px 0; font-size: 16px; font-weight: 600;">Sonraki Adımlar</h3>
                    <ul style="color: #666; line-height: 1.6; margin: 0; padding-left: 20px; font-size: 14px;">
                        <li style="margin-bottom: 8px;"><strong>Takip Kodunuzu Kullanın:</strong> Web sitemizden güncel durumu kontrol edin</li>
                        <li style="margin-bottom: 8px;"><strong>İletişim:</strong> Sorularınız için bizimle iletişime geçebilirsiniz</li>
                    </ul>
                </div>
                
                <!-- Contact Info -->
                <div style="background-color: #e3f2fd; border-radius: 6px; padding: 20px; margin: 20px 0;">
                    <h4 style="color: #1976d2; margin: 0 0 15px 0; font-size: 16px; font-weight: 600;">İletişim Bilgileri</h4>
                    <p style="color: #333; margin: 5px 0; font-size: 14px;">
                        <strong>E-posta:</strong> siparis@mavinefes.com.tr
                    </p>
                    <p style="color: #333; margin: 5px 0; font-size: 14px;">
                        <strong>Telefon:</strong> +90 258 266 55 44
                    </p>
                    <p style="color: #333; margin: 5px 0; font-size: 14px;">
                        <strong>Adres:</strong> Mavi Nefes Yayınları, Zümrüt, Vatan Cd No:240, 20160 Denizli Merkez/Denizli
                    </p>
                </div>
            </div>
            
            <!-- Footer -->
            <div style="background-color: #6c757d; color: white; padding: 20px; text-align: center;">
                <p style="margin: 0 0 5px 0; font-size: 12px; color: rgba(255,255,255,0.8);">
                    Bu e-posta Mavi Nefes Matbaa Takip Sistemi tarafından otomatik olarak gönderilmiştir.
                </p>
                <p style="margin: 0; font-size: 11px; color: rgba(255,255,255,0.6);">
                    © 2025 Mavi Nefes Matbaa. Tüm hakları saklıdır.
                </p>
            </div>
        </div>
    </body>
    </html>
    """
    
    # E-posta gönderme
    if customer_email:
        return send_email_notification(customer_email, subject, body)
    else:
        # Demo için
        demo_email = "demo@example.com"
        return send_email_notification(demo_email, subject, body)



def generate_excel_report(books_data, report_title):
    """Excel raporu oluştur"""
    buffer = BytesIO()
    
    # Workbook oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Kitap Raporu"
    
    # Başlık stilleri
    title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    cell_font = Font(name='Arial', size=11)
    
    # Renkler
    title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    # Kenarlık
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlık
    ws.merge_cells('A1:G1')
    ws['A1'] = report_title
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Tarih
    ws.merge_cells('A2:G2')
    current_time = datetime.now()
    ws['A2'] = f"Rapor Tarihi: {current_time.strftime('%d.%m.%Y %H:%M')}"
    ws['A2'].font = Font(name='Arial', size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Sütun başlıkları
    headers = ['Kitap Adı', 'Yazar', 'Adet', 'Boyut', 'Durum', 'Takip Kodu', 'Müşteri E-posta']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Veri satırları
    if books_data:
        for row, book in enumerate(books_data, 5):
            ws.cell(row=row, column=1, value=book['title']).font = cell_font
            ws.cell(row=row, column=2, value=book['author_name']).font = cell_font
            ws.cell(row=row, column=3, value=book['order_quantity']).font = cell_font
            ws.cell(row=row, column=4, value=book['size']).font = cell_font
            ws.cell(row=row, column=5, value=book['status']).font = cell_font
            ws.cell(row=row, column=6, value=book['track_code']).font = cell_font
            ws.cell(row=row, column=7, value=book['customer_email'] or '').font = cell_font
            
            # Kenarlık ekle
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = thin_border
    else:
        # Veri yoksa mesaj
        ws.merge_cells('A5:G5')
        ws['A5'] = "Bu kategoride raporlanacak veri bulunamadı."
        ws['A5'].font = Font(name='Arial', size=12, italic=True)
        ws['A5'].alignment = Alignment(horizontal='center')
        ws['A5'].border = thin_border
    
    # Sütun genişliklerini ayarla
    column_widths = [30, 20, 10, 15, 15, 15, 25]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Excel dosyasını kaydet
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def login_required(f):
    """Admin giriş kontrolü decorator'ı - İyileştirilmiş güvenlik"""
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            app.logger.warning(f'Yetkisiz erişim denemesi - IP: {request.remote_addr}, Route: {request.endpoint}')
            return redirect(url_for('login'))
        
        # Session timeout kontrolü
        if 'last_activity' in session:
            if datetime.now() - datetime.fromisoformat(session['last_activity']) > app.permanent_session_lifetime:
                session.clear()
                flash('Oturumunuz zaman aşımına uğradı. Lütfen tekrar giriş yapın.', 'warning')
                return redirect(url_for('login'))
        
        session['last_activity'] = datetime.now().isoformat()
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

@app.route('/')
def index():
    """Ana sayfa"""
    return render_template('index.html')

@app.route('/test')
def test():
    """Test endpoint"""
    return jsonify({
        'status': 'ok',
        'message': 'Uygulama çalışıyor',
        'users': 'NARSİST:Mavinefes25'
    })

@app.route('/clear-logout-message', methods=['POST'])
def clear_logout_message():
    """Logout mesajını session'dan temizle"""
    if 'logout_message' in session:
        del session['logout_message']
    return jsonify({'success': True})

@app.route('/health')
def health_check():
    """Sistem sağlık kontrolü"""
    try:
        # Veritabanı bağlantısını test et
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute('SELECT 1')
            db_status = 'healthy'
            conn.close()
        else:
            db_status = 'unhealthy'
        
        # Cache durumunu kontrol et
        try:
            cache.set('health_check', 'ok', timeout=10)
            cache_status = 'healthy'
        except:
            cache_status = 'unhealthy'
        
        # Redis durumunu kontrol et (production)
        redis_status = 'n/a'
        if os.environ.get('FLASK_ENV') == 'production' and os.environ.get('REDIS_URL'):
            try:
                import redis
                r = redis.from_url(os.environ.get('REDIS_URL'))
                r.ping()
                redis_status = 'healthy'
            except:
                redis_status = 'unhealthy'
        
        return jsonify({
            'status': 'healthy',
            'timestamp': datetime.now().isoformat(),
            'database': db_status,
            'cache': cache_status,
            'redis': redis_status,
            'environment': os.environ.get('FLASK_ENV', 'development'),
            'version': '1.0.0'
        })
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

@app.route('/track', methods=['GET', 'POST'])
def track():
    """Kitap takip sayfası"""
    if request.method == 'POST':
        track_code = request.form.get('track_code')
        
        if not track_code:
            flash('Lütfen takip kodunu giriniz.', 'error')
            return render_template('track.html')
        
        conn = get_db_connection()
        if conn:
            try:
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM books WHERE track_code = ?", (track_code,))
                book = cursor.fetchone()
                
                if book:
                    # Veritabanı sütun sırası: id, title, author_name, order_quantity, size, status, track_code, customer_email, created_at, updated_at
                    book_data = {
                        'id': book[0],
                        'title': book[1],
                        'author_name': book[2],
                        'order_quantity': book[3],
                        'size': book[4],
                        'status': book[5],
                        'track_code': book[6],
                        'customer_email': book[7],
                        'created_at': book[8],
                        'updated_at': book[9] if len(book) > 9 else None
                    }
                    
                    app.logger.info(f"Track sayfası - Kitap bulundu: {book_data['title']}, Track kodu: {track_code}, ID: {book_data['id']}")
                    return render_template('track.html', book=book_data, found=True)
                else:
                    flash('Böyle bir kitap bulunamadı.', 'error')
            except Exception as e:
                app.logger.error(f"Track sayfası hatası: {e}")
                flash('Bir hata oluştu.', 'error')
            finally:
                conn.close()
        
        return render_template('track.html')
    
    return render_template('track.html')

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def login():
    """Admin giriş sayfası - İyileştirilmiş güvenlik"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        if not username or not password:
            flash('Kullanıcı adı ve şifre gereklidir.', 'error')
            app.logger.warning(f'Eksik giriş bilgileri - IP: {request.remote_addr}')
            return render_template('login.html')
        
        conn = get_db_connection()
        if conn:
            try:
                cursor = conn.cursor()
                cursor.execute("SELECT id, username, password FROM users WHERE username = ?", (username,))
                user = cursor.fetchone()
                
                if user and verify_password(password, user[2]):
                    session.permanent = True
                    session['admin_logged_in'] = True
                    session['admin_username'] = username
                    session['admin_user_id'] = user[0]
                    
                    # Login başarılı
                    conn.commit()
                    
                    app.logger.info(f'Başarılı giriş - Kullanıcı: {username}, IP: {request.remote_addr}')
                    
                    return redirect(url_for('admin_dashboard'))
                else:
                    flash('Kullanıcı adı veya şifre yanlış.', 'error')
                    app.logger.warning(f'Başarısız giriş denemesi - Kullanıcı: {username}, IP: {request.remote_addr}')
            except Exception as e:
                flash('Giriş sırasında bir hata oluştu.', 'error')
                app.logger.error(f'Giriş hatası: {e}')
            finally:
                conn.close()
        
        return render_template('login.html')
    
    return render_template('login.html')

@app.route('/admin/dashboard')
@login_required
@cache.cached(timeout=60)  # 1 dakika cache
def admin_dashboard():
    """Admin paneli"""
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            
            # Toplam kitap sayısı
            cursor.execute("SELECT COUNT(*) FROM books")
            total_books = cursor.fetchone()[0]
            
            # Aktif siparişler (Hazır olmayanlar)
            cursor.execute("SELECT COUNT(*) FROM books WHERE status != 'Hazır'")
            active_orders = cursor.fetchone()[0]
            
            # Tamamlanan siparişler
            cursor.execute("SELECT COUNT(*) FROM books WHERE status = 'Hazır'")
            completed_orders = cursor.fetchone()[0]
            
            # Son eklenen kitaplar
            cursor.execute("SELECT id, title, author_name, order_quantity, size, status, track_code, customer_email FROM books ORDER BY created_at DESC LIMIT 5")
            recent_books = []
            for row in cursor.fetchall():
                recent_books.append({
                    'id': row[0],
                    'title': row[1],
                    'author_name': row[2],
                    'order_quantity': row[3],
                    'size': row[4],
                    'status': row[5],
                    'track_code': row[6],
                    'customer_email': row[7]
                })
            
            return render_template('admin_dashboard.html', 
                                 total_books=total_books,
                                 active_orders=active_orders,
                                 completed_orders=completed_orders,
                                 recent_books=recent_books)
        except Exception as e:
            flash('Veriler yüklenirken bir hata oluştu.', 'error')
        finally:
            conn.close()
    
    return render_template('admin_dashboard.html')

@app.route('/admin/add', methods=['GET', 'POST'])
@login_required
def add_book():
    """Kitap ekleme sayfası"""
    if request.method == 'POST':
        try:
            # Form verilerini al ve temizle
            title = sanitize_input(request.form.get('title', ''))
            author_name = sanitize_input(request.form.get('author_name', ''))
            track_code = sanitize_input(request.form.get('track_code', ''))
            order_quantity = request.form.get('order_quantity')
            size = sanitize_input(request.form.get('size', ''))
            status = request.form.get('status', 'Sipariş Alındı')
            customer_email = sanitize_input(request.form.get('customer_email', ''))
            send_email = request.form.get('send_email') == 'on'
            
            # Gelişmiş validasyon
            validation_errors = validate_book_data(title, author_name, order_quantity, size)
            if validation_errors:
                for error in validation_errors:
                    flash(error, 'error')
                return render_template('add_book.html')
            
            # E-posta validasyonu
            if send_email and customer_email:
                if not validate_email(customer_email):
                    flash('Geçerli bir e-posta adresi giriniz.', 'error')
                    return render_template('add_book.html')
            
            # Takip kodu oluşturma
            if not track_code:
                track_code = get_unique_track_code()
                app.logger.info(f"Otomatik takip kodu oluşturuldu: {track_code}")
            else:
                # Manuel girilen takip kodunun benzersiz olup olmadığını kontrol et
                if not is_track_code_unique(track_code):
                    flash('Bu takip kodu zaten kullanılıyor. Lütfen farklı bir kod girin.', 'error')
                    return render_template('add_book.html')
            

            
            # Veritabanına kaydet
            conn = get_db_connection()
            if not conn:
                flash('Veritabanı bağlantısı kurulamadı.', 'error')
                return render_template('add_book.html')
            
            try:
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO books (title, author_name, order_quantity, size, status, track_code, customer_email)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (title, author_name, order_quantity, size, status, track_code, customer_email))
                conn.commit()
                
                book_id = cursor.lastrowid
                app.logger.info(f"Kitap başarıyla kaydedildi. ID: {book_id}, Takip Kodu: {track_code}")
                

                
                # Cache'i temizle
                cache.clear()
                
                # E-posta gönderme işlemi
                email_sent = False
                if send_email and customer_email:
                    book_data = {
                        'title': title,
                        'author_name': author_name,
                        'track_code': track_code,
                        'order_quantity': order_quantity,
                        'size': size,
                        'status': status
                    }
                    
                    app.logger.info(f"E-posta gönderiliyor: {customer_email}")
                    email_sent = send_track_code_email(book_data, customer_email)
                    
                    if email_sent:
                        app.logger.info("E-posta başarıyla gönderildi")
                        flash(f'✅ Kitap başarıyla eklendi! Takip kodu: {track_code} - E-posta gönderildi.', 'success')
                    else:
                        app.logger.warning("E-posta gönderilemedi")
                        flash(f'⚠️ Kitap başarıyla eklendi (Takip kodu: {track_code}) ancak e-posta gönderilemedi.', 'warning')
                else:
                    flash(f'✅ Kitap başarıyla eklendi! Takip kodu: {track_code}', 'success')
                
                # SocketIO bildirimi
                socketio.emit('book_added', {
                    'book_id': book_id,
                    'title': title,
                    'track_code': track_code,
                    'message': f'Yeni kitap eklendi: {title}'
                })
                
                return redirect(url_for('admin_dashboard'))
                
            except Exception as e:
                print(f"Veritabanı hatası: {e}")
                flash('Kitap eklenirken bir hata oluştu.', 'error')
                return render_template('add_book.html')
            finally:
                conn.close()
                
        except Exception as e:
            app.logger.error(f"Genel hata: {e}")
            flash('Beklenmeyen bir hata oluştu.', 'error')
            return render_template('add_book.html')
    
    return render_template('add_book.html')

@app.route('/admin/update/<int:book_id>', methods=['GET', 'POST'])
@login_required
def update_book(book_id):
    """Kitap güncelleme sayfası"""
    if request.method == 'POST':
        title = request.form.get('title')
        author_name = request.form.get('author_name')
        order_quantity = request.form.get('order_quantity')
        size = request.form.get('size')
        status = request.form.get('status')
        customer_email = request.form.get('customer_email')
        send_status_email = request.form.get('send_status_email') == 'on'
        

        
        conn = get_db_connection()
        if not conn:
            flash('Veritabanı bağlantısı kurulamadı.', 'error')
            return redirect(url_for('admin_dashboard'))
        
        try:
            cursor = conn.cursor()
            
            # Önce mevcut kitap bilgilerini al
            cursor.execute("SELECT * FROM books WHERE id = ?", (book_id,))
            book = cursor.fetchone()
            
            if not book:
                flash('Kitap bulunamadı.', 'error')
                return redirect(url_for('admin_dashboard'))
            
            old_status = book[5]  # Eski durum
            old_customer_email = book[7]  # Eski e-posta
            
            # Güncelleme işlemi - İyileştirilmiş
            cursor.execute("""
                UPDATE books SET title=?, author_name=?, order_quantity=?, size=?, status=?, customer_email=?, updated_at=CURRENT_TIMESTAMP
                    WHERE id=?
            """, (title, author_name, order_quantity, size, status, customer_email, book_id))
            
            conn.commit()
            

            
            # Cache'i temizle
            cache.clear()
            
            # Durum değişikliği kontrolü ve e-posta gönderme
            if send_status_email and customer_email and status != old_status:
                book_data = {
                    'title': title,
                    'author_name': author_name,
                    'track_code': book[6],  # track_code sütunu
                    'order_quantity': order_quantity,
                    'size': size
                }
                
                email_sent = send_status_update_email(book_data, status, customer_email)
                if email_sent:
                    flash('Kitap başarıyla güncellendi ve durum güncellemesi e-postası gönderildi.', 'success')
                else:
                    flash('Kitap başarıyla güncellendi ancak e-posta gönderilemedi.', 'warning')
            else:
                flash('Kitap başarıyla güncellendi.', 'success')
            
            # SocketIO bildirimi
            socketio.emit('status_updated', {
                'book_id': book_id,
                'new_status': status,
                'message': f'Kitap durumu güncellendi: {status}'
            })
            
            return redirect(url_for('admin_dashboard'))
        except Exception as e:
            flash('Kitap güncellenirken bir hata oluştu.', 'error')
            app.logger.error(f"Güncelleme hatası: {e}")
        finally:
            conn.close()
    
    # GET isteği - Kitap bilgilerini getir
    conn = get_db_connection()
    if not conn:
        flash('Veritabanı bağlantısı kurulamadı.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM books WHERE id = ?", (book_id,))
        book = cursor.fetchone()
        
        if book:
            # Veritabanı sütun sırası: id, title, author_name, order_quantity, size, status, track_code, customer_email, created_at, updated_at
            book_data = {
                'id': book[0],
                'title': book[1],
                'author_name': book[2],
                'order_quantity': book[3],
                'size': book[4],
                'status': book[5],
                'track_code': book[6],
                'customer_email': book[7],
                'created_at': book[8],
                'updated_at': book[9] if len(book) > 9 else None
            }
            app.logger.info(f"Kitap verisi yüklendi - ID: {book[0]}")
            return render_template('update_book.html', book=book_data)
        else:
            flash('Kitap bulunamadı.', 'error')
            return redirect(url_for('admin_dashboard'))
    except Exception as e:
        flash('Kitap bilgileri yüklenirken bir hata oluştu.', 'error')
        app.logger.error(f"Veri yükleme hatası: {e}")
        return redirect(url_for('admin_dashboard'))
    finally:
        conn.close()

@app.route('/admin/delete/<int:book_id>', methods=['DELETE'])
@login_required
def delete_book(book_id):
    """Kitap silme endpoint'i"""
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            
            # Önce kitabın var olup olmadığını kontrol et
            cursor.execute("SELECT * FROM books WHERE id = ?", (book_id,))
            book = cursor.fetchone()
            
            if not book:
                return jsonify({'success': False, 'error': 'Kitap bulunamadı'})
            
            # Görsel dosyası sütunu yok, bu kısmı kaldırıyoruz
            
            # Kitabı veritabanından sil
            cursor.execute("DELETE FROM books WHERE id = ?", (book_id,))
            conn.commit()
            
            return jsonify({'success': True, 'message': 'Kitap başarıyla silindi'})
            
        except Exception as e:
            app.logger.error(f"Kitap silme hatası: {e}")
            return jsonify({'success': False, 'error': str(e)})
        finally:
            conn.close()
    else:
        return jsonify({'success': False, 'error': 'Veritabanı bağlantısı kurulamadı'})

@app.route('/logout')
def logout():
    """Oturumu kapat - İyileştirilmiş"""
    username = session.get('admin_username', 'Unknown')
    user_id = session.get('admin_user_id')
    
    app.logger.info(f'Kullanıcı çıkış yaptı - Kullanıcı: {username}, IP: {request.remote_addr}')
    session.clear()
    
    # Flash mesajını session'a kaydet ve index'e yönlendir
    session['logout_message'] = 'Başarıyla çıkış yaptınız.'
    return redirect(url_for('index'))

@app.route('/admin/books/all')
@login_required
def get_all_books():
    """Tüm kitapları JSON formatında döndür - Sayfalama ile"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    # Güvenlik: per_page limitini kontrol et
    per_page = min(per_page, 100)
    
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            
            # Toplam kayıt sayısı
            cursor.execute("SELECT COUNT(*) FROM books")
            total = cursor.fetchone()[0]
            
            # Sayfalama hesaplamaları
            offset = (page - 1) * per_page
            
            cursor.execute("""
                SELECT id, title, author_name, track_code, order_quantity, size, status, customer_email, created_at
                FROM books 
                ORDER BY created_at DESC
                LIMIT ? OFFSET ?
            """, (per_page, offset))
            
            books = []
            for row in cursor.fetchall():
                books.append({
                    'id': row[0],
                    'title': row[1],
                    'author_name': row[2],
                    'track_code': row[3],
                    'order_quantity': row[4],
                    'size': row[5],
                    'status': row[6],
                    'customer_email': row[7],
                    'created_at': row[8].isoformat() if row[8] else None
                })
            
            return jsonify({
                'books': books,
                'pagination': {
                    'page': page,
                    'per_page': per_page,
                    'total': total,
                    'pages': (total + per_page - 1) // per_page
                }
            })
        except Exception as e:
            app.logger.error(f"Veritabanı hatası: {e}")
            return jsonify({'error': 'Veritabanı hatası'}), 500
        finally:
            conn.close()
    return jsonify({'error': 'Veritabanı bağlantı hatası'}), 500

@app.route('/admin/books/active')
@login_required
def get_active_books():
    """Aktif siparişleri JSON formatında döndür"""
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, title, author_name, track_code, order_quantity, size, status, customer_email, created_at
                FROM books 
                WHERE status != 'Hazır'
                ORDER BY created_at DESC
            """)
            books = []
            for row in cursor.fetchall():
                books.append({
                    'id': row[0],
                    'title': row[1],
                    'author_name': row[2],
                    'track_code': row[3],
                    'order_quantity': row[4],
                    'size': row[5],
                    'status': row[6],
                    'customer_email': row[7],
                    'created_at': row[8].isoformat() if row[8] else None
                })
            return jsonify({'books': books})
        except Exception as e:
            app.logger.error(f"Veritabanı hatası: {e}")
            return jsonify({'error': 'Veritabanı hatası'}), 500
        finally:
            conn.close()
    return jsonify({'error': 'Veritabanı bağlantı hatası'}), 500

@app.route('/admin/books/completed')
@login_required
def get_completed_books():
    """Tamamlanan siparişleri JSON formatında döndür"""
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, title, author_name, track_code, order_quantity, size, status, customer_email, created_at
                FROM books 
                WHERE status = 'Hazır'
                ORDER BY created_at DESC
            """)
            books = []
            for row in cursor.fetchall():
                books.append({
                    'id': row[0],
                    'title': row[1],
                    'author_name': row[2],
                    'track_code': row[3],
                    'order_quantity': row[4],
                    'size': row[5],
                    'status': row[6],
                    'customer_email': row[7],
                    'created_at': row[8].isoformat() if row[8] else None
                })
            return jsonify({'books': books})
        except Exception as e:
            app.logger.error(f"Veritabanı hatası: {e}")
            return jsonify({'error': 'Veritabanı hatası'}), 500
        finally:
            conn.close()
    return jsonify({'error': 'Veritabanı bağlantı hatası'}), 500

@app.route('/admin/books/all-page')
@login_required
def all_books_page():
    """Tüm kitaplar sayfası"""
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT id, title, author_name, order_quantity, size, status, track_code, customer_email FROM books ORDER BY created_at DESC")
            books = []
            for row in cursor.fetchall():
                books.append({
                    'id': row[0],
                    'title': row[1],
                    'author_name': row[2],
                    'order_quantity': row[3],
                    'size': row[4],
                    'status': row[5],
                    'track_code': row[6],
                    'customer_email': row[7]
                })
            return render_template('books_list.html', books=books, title="Tüm Kitaplar", type="all")
        except Exception as e:
            flash('Veriler yüklenirken bir hata oluştu.', 'error')
        finally:
            conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/books/active-page')
@login_required
def active_books_page():
    """Aktif siparişler sayfası"""
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT id, title, author_name, order_quantity, size, status, track_code, customer_email FROM books WHERE status != 'Hazır' ORDER BY created_at DESC")
            books = []
            for row in cursor.fetchall():
                app.logger.debug(f"Veritabanı sütun sırası: {list(row)}")
                books.append({
                    'id': row[0],
                    'title': row[1],
                    'author_name': row[2],
                    'order_quantity': row[3],
                    'size': row[4],
                    'status': row[5],
                    'track_code': row[6],
                    'customer_email': row[7]
                })
            return render_template('books_list.html', books=books, title="Aktif Siparişler", type="active")
        except Exception as e:
            flash('Veriler yüklenirken bir hata oluştu.', 'error')
        finally:
            conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/books/completed-page')
@login_required
def completed_books_page():
    """Tamamlanan siparişler sayfası"""
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT id, title, author_name, order_quantity, size, status, track_code, customer_email FROM books WHERE status = 'Hazır' ORDER BY created_at DESC")
            books = []
            for row in cursor.fetchall():
                books.append({
                    'id': row[0],
                    'title': row[1],
                    'author_name': row[2],
                    'order_quantity': row[3],
                    'size': row[4],
                    'status': row[5],
                    'track_code': row[6],
                    'customer_email': row[7]
                })
            return render_template('books_list.html', books=books, title="Tamamlanan Siparişler", type="completed")
        except Exception as e:
            flash('Veriler yüklenirken bir hata oluştu.', 'error')
        finally:
            conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/generate-track-code')
@login_required
def generate_track_code_api():
    """API endpoint for generating unique track code"""
    try:
        # Rate limiting kontrolü (opsiyonel)
        track_code = get_unique_track_code()
        
        # Log kaydı
        app.logger.info(f"Takip kodu oluşturuldu: {track_code} - Kullanıcı: {session.get('admin_username', 'Unknown')}")
        
        return jsonify({
            'success': True, 
            'track_code': track_code,
            'message': 'Takip kodu başarıyla oluşturuldu'
        })
    except Exception as e:
        app.logger.error(f"Takip kodu oluşturma hatası: {e}")
        return jsonify({
            'success': False, 
            'error': 'Takip kodu oluşturulurken bir hata oluştu'
        }), 500

@app.route('/admin/export-excel/<report_type>')
@login_required
def export_excel_report(report_type):
    """Excel rapor indir"""
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            
            if report_type == 'all':
                cursor.execute("""
                    SELECT id, title, author_name, order_quantity, size, status, track_code, customer_email, created_at
                    FROM books ORDER BY created_at DESC
                """)
            elif report_type == 'active':
                cursor.execute("""
                    SELECT id, title, author_name, order_quantity, size, status, track_code, customer_email, created_at
                    FROM books WHERE status IN ('Sipariş Alındı', 'Hazırlanıyor', 'Üretimde') ORDER BY created_at DESC
                """)
            elif report_type == 'completed':
                cursor.execute("""
                    SELECT id, title, author_name, order_quantity, size, status, track_code, customer_email, created_at
                    FROM books WHERE status = 'Hazır' ORDER BY created_at DESC
                """)
            else:
                return jsonify({'error': 'Geçersiz rapor türü'}), 400
            
            books = []
            for row in cursor.fetchall():
                books.append({
                    'id': row[0],
                    'title': row[1],
                    'author_name': row[2],
                    'order_quantity': row[3],
                    'size': row[4],
                    'status': row[5],
                    'track_code': row[6],
                    'customer_email': row[7] if row[7] else '',
                    'created_at': row[8] if row[8] else ''
                })
            
            # Excel oluştur
            excel_buffer = generate_excel_report(books, f"Matbaa Raporu - {report_type.title()}")
            
            # Dosya adı oluştur
            current_time = datetime.now()
            timestamp = current_time.strftime('%Y%m%d_%H%M%S')
            filename = f"matbaa_rapor_{report_type}_{timestamp}.xlsx"
            
            excel_buffer.seek(0)
            return send_file(
                excel_buffer,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except Exception as e:
            return jsonify({'error': str(e)}), 500
        finally:
            conn.close()
    
    return jsonify({'error': 'Veritabanı bağlantı hatası'}), 500





@app.route('/contact', methods=['POST'])
@limiter.limit("3 per minute")
def contact():
    """İletişim formu gönderimi - Rate limiting ile"""
    try:
        name = request.form.get('name')
        email = request.form.get('email')
        message = request.form.get('message')
        
        if not name or not email or not message:
            return jsonify({'success': False, 'message': 'Tüm alanları doldurun'}), 400
        
        conn = get_db_connection()
        if conn:
            try:
                cursor = conn.cursor()
                cursor.execute(
                    'INSERT INTO contact_messages (name, email, message) VALUES (?, ?, ?)',
                    (name, email, message)
                )
                conn.commit()
                
                # E-posta bildirimi gönder (opsiyonel)
                if EMAIL_ENABLED:
                    try:
                        subject = f"Yeni İletişim Formu Mesajı - {name}"
                        body = f"""
                        Yeni bir iletişim formu mesajı alındı:
                        
                        Ad Soyad: {name}
                        E-posta: {email}
                        Mesaj: {message}
                        
                        Tarih: {datetime.now().strftime('%d.%m.%Y %H:%M')}
                        """
                        
                        # Admin e-posta adresine bildirim gönder
                        admin_email = 'siparis@mavinefes.com.tr'
                        send_email_notification(admin_email, subject, body)
                        
                    except Exception as e:
                        app.logger.error(f"E-posta gönderme hatası: {e}")
                
                return jsonify({'success': True, 'message': 'Mesajınız başarıyla gönderildi'}), 200
                
            except Exception as e:
                return jsonify({'success': False, 'message': f'Veritabanı hatası: {str(e)}'}), 500
            finally:
                conn.close()
        
        return jsonify({'success': False, 'message': 'Veritabanı bağlantı hatası'}), 500
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Sistem hatası: {str(e)}'}), 500

@app.route('/admin/contact-messages')
@login_required
def contact_messages():
    """İletişim mesajları sayfası"""
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT id, name, email, message, is_read, created_at 
                FROM contact_messages 
                ORDER BY created_at DESC
            ''')
            
            messages = []
            for row in cursor.fetchall():
                # created_at alanını datetime objesi olarak parse et
                created_at = row[5]
                if isinstance(created_at, str):
                    try:
                        from datetime import datetime
                        created_at = datetime.strptime(created_at, '%Y-%m-%d %H:%M:%S')
                    except:
                        created_at = created_at
                
                messages.append({
                    'id': row[0],
                    'name': row[1],
                    'email': row[2],
                    'message': row[3],
                    'is_read': bool(row[4]),
                    'created_at': created_at
                })
            
            return render_template('contact_messages.html', contact_messages=messages)
            
        except Exception as e:
            flash(f'Mesajlar yüklenirken hata oluştu: {str(e)}', 'error')
            return render_template('contact_messages.html', contact_messages=[])
        finally:
            conn.close()
    
    flash('Veritabanı bağlantı hatası', 'error')
    return render_template('contact_messages.html', contact_messages=[])

@app.route('/admin/contact-messages/<int:message_id>')
@login_required
def get_contact_message(message_id):
    """Tekil mesaj detayını getir"""
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT id, name, email, message, is_read, created_at 
                FROM contact_messages 
                WHERE id = ?
            ''', (message_id,))
            
            row = cursor.fetchone()
            if row:
                # created_at alanını datetime objesi olarak parse et
                created_at = row[5]
                if isinstance(created_at, str):
                    try:
                        from datetime import datetime
                        created_at = datetime.strptime(created_at, '%Y-%m-%d %H:%M:%S')
                        created_at = created_at.strftime('%d.%m.%Y %H:%M')
                    except:
                        created_at = created_at
                elif created_at:
                    created_at = created_at.strftime('%d.%m.%Y %H:%M')
                else:
                    created_at = ''
                
                message = {
                    'id': row[0],
                    'name': row[1],
                    'email': row[2],
                    'message': row[3],
                    'is_read': bool(row[4]),
                    'created_at': created_at
                }
                return jsonify({'success': True, 'message': message})
            else:
                return jsonify({'success': False, 'error': 'Mesaj bulunamadı'}), 404
                
        except Exception as e:
            return jsonify({'success': False, 'error': f'Veritabanı hatası: {str(e)}'}), 500
        finally:
            conn.close()
    
    return jsonify({'success': False, 'error': 'Veritabanı bağlantı hatası'}), 500

@app.route('/admin/contact-messages/<int:message_id>/read', methods=['POST'])
@login_required
def mark_message_read(message_id):
    """Mesajı okundu olarak işaretle"""
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute('UPDATE contact_messages SET is_read = 1 WHERE id = ?', (message_id,))
            conn.commit()
            
            return jsonify({'success': True})
                
        except Exception as e:
            return jsonify({'success': False, 'error': f'Veritabanı hatası: {str(e)}'}), 500
        finally:
            conn.close()
    
    return jsonify({'success': False, 'error': 'Veritabanı bağlantı hatası'}), 500

@app.route('/admin/contact-messages/<int:message_id>', methods=['DELETE'])
@login_required
def delete_contact_message(message_id):
    """Tekil mesaj silme"""
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM contact_messages WHERE id = ?', (message_id,))
            conn.commit()
            
            if cursor.rowcount > 0:
                return jsonify({'success': True})
            else:
                return jsonify({'error': 'Mesaj bulunamadı'}), 404
                
        except Exception as e:
            return jsonify({'error': f'Veritabanı hatası: {str(e)}'}), 500
        finally:
            conn.close()
    
    return jsonify({'error': 'Veritabanı bağlantı hatası'}), 500

@app.route('/admin/contact-messages/mark-all-read', methods=['POST'])
@login_required
def mark_all_messages_read():
    """Tüm mesajları okundu olarak işaretle"""
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute('UPDATE contact_messages SET is_read = 1 WHERE is_read = 0')
            conn.commit()
            
            return jsonify({'success': True, 'updated_count': cursor.rowcount})
                
        except Exception as e:
            return jsonify({'error': f'Veritabanı hatası: {str(e)}'}), 500
        finally:
            conn.close()
    
    return jsonify({'error': 'Veritabanı bağlantı hatası'}), 500

@app.route('/admin/contact-messages/delete-all', methods=['DELETE'])
@login_required
def delete_all_messages():
    """Tüm mesajları sil"""
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM contact_messages')
            conn.commit()
            
            return jsonify({'success': True, 'deleted_count': cursor.rowcount})
                
        except Exception as e:
            return jsonify({'error': f'Veritabanı hatası: {str(e)}'}), 500
        finally:
            conn.close()
    
    return jsonify({'error': 'Veritabanı bağlantı hatası'}), 500









# SocketIO event handlers
@socketio.on('connect')
def handle_connect(auth=None):
    app.logger.info('Client connected')

@socketio.on('disconnect')
def handle_disconnect():
    app.logger.info('Client disconnected')

@socketio.on('status_update')
def handle_status_update(data):
    """Durum güncellemesi bildirimi"""
    emit('status_updated', data, broadcast=True)

# Error Handlers
@app.errorhandler(404)
def not_found_error(error):
    """404 Hata sayfası"""
    app.logger.warning(f'404 hatası - IP: {request.remote_addr}, URL: {request.url}')
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    """500 Hata sayfası"""
    app.logger.error(f'500 hatası - IP: {request.remote_addr}, URL: {request.url}, Error: {error}')
    
    # Production'da detaylı hata bilgisi gösterme
    if os.environ.get('FLASK_ENV') == 'production':
        return render_template('errors/500.html'), 500
    else:
        # Development'ta detaylı hata
        return f'<h1>Internal Server Error</h1><p>{error}</p>', 500

@app.errorhandler(403)
def forbidden_error(error):
    """403 Hata sayfası"""
    app.logger.warning(f'403 hatası - IP: {request.remote_addr}, URL: {request.url}')
    return render_template('errors/403.html'), 403

@app.errorhandler(429)
def ratelimit_handler(e):
    """Rate limit aşıldığında"""
    app.logger.warning(f'Rate limit aşıldı - IP: {request.remote_addr}, URL: {request.url}')
    return jsonify({'error': 'Çok fazla istek gönderdiniz. Lütfen bekleyin.'}), 429



if __name__ == '__main__':
    try:
        # Güvenlik kontrolü
        if not os.environ.get('SECRET_KEY'):
            print("❌ SECRET_KEY environment variable bulunamadı!")
            print("💡 .env dosyası oluşturun veya SECRET_KEY ayarlayın")
            sys.exit(1)
        
        if not os.environ.get('MAIL_PASSWORD') and os.environ.get('EMAIL_ENABLED', 'True').lower() == 'true':
            print("⚠️  MAIL_PASSWORD bulunamadı - Email sistemi devre dışı")
            os.environ['EMAIL_ENABLED'] = 'False'
        
        init_db() # Veritabanını başlat
        print("✅ Veritabanı başarıyla başlatıldı")
        
        # Production/Development port ayarı
        port = int(os.environ.get('PORT', 8080))
        debug = os.environ.get('FLASK_ENV') != 'production'
        
        print(f"🚀 Uygulama başlatılıyor - Port: {port}, Debug: {debug}")
        socketio.run(app, host='0.0.0.0', port=port, debug=debug)
        
    except Exception as e:
        print(f"❌ Uygulama başlatılamadı: {e}")
        sys.exit(1) 